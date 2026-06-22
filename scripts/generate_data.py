#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""从 01-Raw-Materials/01-价格表 读取 Excel，生成 data.js 和 thor-pricing.json5"""

from __future__ import annotations

import glob
import json
import os
import re
import sys
from datetime import date

try:
    from openpyxl import load_workbook
except ImportError:
    print("请先安装: pip install openpyxl")
    sys.exit(1)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW_DIR = os.path.join(ROOT, "01-Raw-Materials", "01-价格表")
OUT_JS = os.path.join(ROOT, "data.js")
OUT_JSON5 = os.path.join(ROOT, "02-Database", "thor-pricing.json5")

CABINET_TYPES = ["洗手盘地柜", "地柜", "灶台吊柜", "吊柜", "高柜", "上翻门吊柜"]
THICKNESSES = ["5mm", "6mm", "8mm", "10mm", "12mm", "15mm", "19mm", "39mm"]

# 报价单使用的 55 项热门五金（名称与 五金.xlsx 对齐）
HARDWARE_NAMES = [
    "百隆107°集成阻尼全盖铰链",
    "百隆107°集成阻尼半盖铰链",
    "百隆107°集成阻尼内掩铰链",
    "百隆95°集成阻尼转角柜门铰链",
    "反弹器",
    "百隆155°集成阻尼全盖铰链",
    "天地铰链(铝框)",
    "百隆魅宝低抽400",
    "百隆魅宝扶杆高抽400",
    "百隆全拉托底轨410(阻尼)",
    "百隆反弹全拉托底轨410",
    "百隆乐薄低抽450",
    "百隆乐薄高抽450",
    "百隆乐薄内低抽450",
    "百隆全拉隔板锁定装置",
    "铝合金内分隔件600",
    "铝合金内分隔件800",
    "可调实木抽屉分隔件A",
    "希勒3.0三边平篮800",
    "希勒3.0三边碗碟篮800",
    "希勒3.0调味篮400",
    "希勒4.0三边平篮800",
    "希勒4.0三边碗碟篮800",
    "希勒飞碟篮900",
    "希勒小怪物篮900",
    "希勒四层高柜连动拉篮450",
    "希勒六层高柜连动拉篮600",
    "宁卡抽拉式垃圾桶",
    "L型免拉手",
    "U型免拉手",
    "L型灯光免拉手",
    "水晶拉手(32孔距)",
    "水晶拉手(96孔距)",
    "金属拉手(96孔距)",
    "金属拉手(128孔距)",
    "皮革拉手",
    "嵌入式灯条600",
    "嵌入式灯条800",
    "嵌入式斜照灯条600",
    "抽屉感应灯600",
    "抽屉感应灯800",
    "双面发光层板灯600",
    "双面发光层板灯800",
    "嵌入式手扫感应灯600",
    "单门碰感应开关",
    "双门碰感应开关",
    "手扫感应开关",
    "触摸感应开关",
    "人体感应开关",
    "集控电源60W",
    "集控电源30W",
    "级联线",
    "玻璃夹板灯600",
    "嵌入式灯条1500",
    "嵌入式灯条1800",
]

RULES = {
    "countertop": {
        "standard": {
            "width": 600,
            "thickness": 15,
            "frontHeight": "12/19/39",
            "splashHeight": 46,
            "backHeight": 50,
        },
        "widthMultiplier": [
            {"max": 600, "multiplier": 1.0},
            {"max": 650, "multiplier": 1.3},
            {"max": 900, "multiplier": 1.5},
            {"max": 1050, "multiplier": 2.0},
        ],
        "minLength": 1000,
        "edgeUpcharge": {"前斜边": 500, "前弧边": 500, "unit": "元/米"},
    },
    "processing": {
        "stoveHole": 300,
        "sinkHole": 900,
        "sinkHoleNote": "公司配套水槽可免收",
        "woodCrate": 300,
        "unit": "元/个",
    },
    "cabinet": {
        "included": "柜体及连接件、门铰、门板、可调脚、柜体铝合金免拉手",
        "wallCabinet": {"halfPrice": 0.65, "halfHeightRatio": 0.5},
        "tallCabinet": {"halfMultiplier": 2, "fullMultiplier": 3},
        "topPanel": {
            "noPattern": {"threshold": 100, "singleLayer": 400, "unit": "元/米"},
            "minArea": 0.1,
        },
        "baseShoe": {"布纹板": 270, "喷粉": 400, "unit": "元/米"},
        "doorUpcharges": {
            "极光M.L48": 200,
            "极光M.L49": 120,
            "斜边门": 400,
            "弧边门": 400,
            "不锈钢压型/铝框造型": 1000,
            "格栅门": 2000,
            "unit": "元/㎡",
        },
        "thicknessUpcharges": {
            "18mm": {"default": True},
            "22mm": {"upcharge": 1000, "withHandle": 1500, "unit": "元/㎡"},
        },
        "materialMultiplier": {
            "哑光小雨滴板": 1.0,
            "布纹板": 1.0,
            "喷粉柜体": 1.1,
            "云纹板体": 1.05,
        },
        "doubleBackPanel": {"upcharge": 1150, "unit": "元/㎡"},
        "wallHandle": {
            "default": "斜边门型或门板下扣10mm",
            "bottomHandleUpcharge": 150,
            "unit": "元/m",
        },
    },
}


def find_file(*patterns: str) -> str:
    if not os.path.isdir(RAW_DIR):
        raise FileNotFoundError(f"目录不存在: {RAW_DIR}")
    for pat in patterns:
        matches = [
            os.path.join(RAW_DIR, f)
            for f in os.listdir(RAW_DIR)
            if not f.startswith("~$") and re.search(pat, f, re.I)
        ]
        if matches:
            return sorted(matches, key=os.path.getmtime, reverse=True)[0]
    raise FileNotFoundError(f"在 {RAW_DIR} 未找到匹配文件: {patterns}")


SHORT_SERIES = {"云纹", "韵纹", "臻纹", "绚丽", "金属", "玉润"}


def norm_series(raw) -> str | None:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or s == "台面系列":
        return None
    if "质影" in s and "原色" in s:
        return "质影/原色"
    s = re.sub(r"系列$", "", s)
    if s == "铂金":
        return "铂金系列"
    if s in SHORT_SERIES:
        return s + "系列"
    return s.strip()


def spec_key(raw) -> str | None:
    if not raw:
        return None
    text = (
        str(raw)
        .replace("（", "")
        .replace("）", "")
        .replace("(", "")
        .replace(")", "")
        .replace("高", "")
        .replace(" ", "")
    )
    m = re.search(r"(地柜|吊柜).*?(\d+)", text)
    if m:
        return m.group(1) + m.group(2)
    return None


def norm_countertop_series(raw: str) -> str:
    s = raw.strip()
    if s.startswith("琉-"):
        s = "琉晶" + s[1:]
    if "/" in s and "GY" in s and len(s) > 40:
        return "琉晶等实心多层"
    if s.startswith("304#"):
        return "304#喷砂纹实心"
    if "晶砾" in s or "GY11" in s:
        return "晶砾-银砂GY11"
    if "GY01" in s:
        return "琉晶-丝纹GY01"
    if "GY02" in s:
        return "琉晶-绘纹GY02"
    if "GY03" in s:
        return "晶灵-蚀刻GY03"
    if "GY04" in s:
        return "晶灵-雾砂GY04"
    if "GY05" in s:
        return "晶灵-珠星GY05"
    if "GY08" in s:
        return "晶灵-雾瓷GY08"
    if "GY09" in s:
        return "品灵-柳叶GY09"
    if "GY10" in s:
        return "晶灵-印迹GY10"
    return s


def norm_thickness(raw) -> str | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        n = int(raw)
        return f"{n}mm"
    s = str(raw).strip().lower().replace(" ", "")
    if not s:
        return None
    if s.endswith("mm"):
        return s
    if re.fullmatch(r"\d+", s):
        return f"{s}mm"
    return None


def hw_icon(name: str, category: str = "") -> str:
    n = name + category
    if any(k in n for k in ("铰链", "反弹")):
        return "hinge"
    if any(k in n for k in ("抽", "轨", "分隔")):
        return "drawer"
    if any(k in n for k in ("篮", "垃圾桶", "飞碟", "怪物")):
        return "basket"
    if any(k in n for k in ("拉手", "免拉手")):
        return "handle"
    if any(k in n for k in ("灯", "感应灯")):
        return "light"
    if any(k in n for k in ("开关", "电源")):
        return "switch"
    if "线" in n:
        return "wire"
    return "hinge"


def normalize_hw_name(name: str) -> str:
    return re.sub(r"[\s（）()\-°]", "", name)


def find_hardware(catalog: dict[str, dict], target: str) -> dict | None:
    if target in catalog:
        return catalog[target]
    compact = normalize_hw_name(target)
    for k, v in catalog.items():
        nk = normalize_hw_name(k)
        if nk == compact or compact in nk or nk in compact:
            return v
    # 左开/右开变体：希勒飞碟篮900 → 希勒飞碟篮左开900
    if "900" in target and ("飞碟" in target or "怪物" in target):
        for side in ("左开", "右开", "左", "右"):
            variant = target.replace("篮900", f"篮{side}900")
            if variant in catalog:
                return catalog[variant]
            for k, v in catalog.items():
                if variant.replace(" ", "") in k.replace(" ", ""):
                    return v
    key = re.sub(r"[^\u4e00-\u9fffA-Za-z0-9]", "", target)[:8]
    for k, v in catalog.items():
        if key and key in normalize_hw_name(k):
            return v
    return None


def hw_unit(unit_raw) -> str:
    if not unit_raw:
        return "个"
    s = str(unit_raw)
    if "米" in s:
        return "米"
    if "套" in s:
        return "套"
    return "个"


def parse_cabinet_and_countertop(path: str) -> tuple[dict, dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[1]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    pricing: dict[str, dict] = {}
    for row in rows[2:]:
        series_raw, spec_raw, price, panel = (
            row[0] if len(row) > 0 else None,
            row[2] if len(row) > 2 else None,
            row[7] if len(row) > 7 else None,
            row[8] if len(row) > 8 else None,
        )
        if series_raw and "台面" in str(series_raw):
            break
        series = norm_series(series_raw)
        key = spec_key(spec_raw)
        if not series or not key or price is None:
            continue
        try:
            pricing.setdefault(series, {})[key] = round(float(price))
            if panel is not None:
                pricing[series]["panel"] = round(float(panel))
        except (TypeError, ValueError):
            continue

    ct_pricing: dict[str, dict] = {}
    current: str | None = None
    started = False
    for row in rows:
        a = row[0] if len(row) > 0 else None
        c = row[2] if len(row) > 2 else None
        d = row[3] if len(row) > 3 else None
        if a and "台面" in str(a):
            started = True
            continue
        if not started:
            continue
        if a and str(a).strip():
            text = str(a).strip()
            if "GY" in text or text.startswith("304#") or "实心" in text:
                current = norm_countertop_series(text)
                ct_pricing.setdefault(current, {t: 0 for t in THICKNESSES})
        if current and c is not None and d is not None:
            thick = norm_thickness(c)
            if not thick:
                continue
            try:
                ct_pricing.setdefault(current, {t: 0 for t in THICKNESSES})[thick] = round(float(d))
            except (TypeError, ValueError):
                continue

    return pricing, ct_pricing


def parse_hardware(path: str) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    catalog: dict[str, dict] = {}
    for row in rows[1:]:
        if len(row) < 8:
            continue
        category = str(row[1] or "").strip()
        name = str(row[2] or "").strip().replace("\n", "")
        price = row[7]
        if not name or price is None:
            continue
        try:
            price_f = round(float(price))
        except (TypeError, ValueError):
            continue
        catalog[name] = {
            "name": name,
            "price": price_f,
            "unit": hw_unit(row[6]),
            "category": category,
        }

    hardware = []
    for i, target in enumerate(HARDWARE_NAMES, 1):
        item = find_hardware(catalog, target)
        if not item:
            print(f"  [warn] 五金未找到: {target}")
            item = {"name": target, "price": 0, "unit": "个", "category": ""}
        hardware.append(
            {
                "id": f"h{i}",
                "name": item["name"],
                "price": item["price"],
                "unit": item["unit"],
                "icon": hw_icon(item["name"], item.get("category", "")),
            }
        )
    return hardware


def build_data(cabinet_pricing: dict, ct_pricing: dict, hardware: list[dict]) -> dict:
    series_order = list(cabinet_pricing.keys())
    ct_series = list(ct_pricing.keys())
    return {
        "cabinet": {
            "series": series_order,
            "types": CABINET_TYPES,
            "pricing": cabinet_pricing,
            "images": {},
        },
        "countertop": {
            "series": ct_series,
            "thicknesses": THICKNESSES,
            "pricing": ct_pricing,
            "images": {},
        },
        "hardware": hardware,
        "rules": RULES,
    }


def to_js(data: dict, sources: list[str]) -> str:
    src = " + ".join(os.path.basename(s) for s in sources)
    body = json.dumps(data, ensure_ascii=False, indent=2)
    return (
        f"// THOR 报价数据库 - 由 01-Raw-Materials/01-价格表 自动生成\n"
        f"// 源文件: {src}\n"
        f"// 生成日期: {date.today().isoformat()}\n"
        f"const THOR_DATA = {body};\n"
    )


def to_json5(data: dict, sources: list[str]) -> str:
    src = " + ".join(os.path.basename(s) for s in sources)
    lines = [
        "// THOR 索而橱柜报价系统 - 产品数据库",
        f"// 源数据来自: {src}",
        f"// 最后更新: {date.today().isoformat()}",
        "",
        json.dumps(data, ensure_ascii=False, indent=2),
        "",
    ]
    return "\n".join(lines)


def main() -> int:
    print(f"[THOR] 读取源文件: {RAW_DIR}")
    cab_path = find_file(r"2025.*\.xlsx$", r"THOR.*\.xlsx$")
    hw_path = find_file(r"五金\.xlsx$", r"^((?!2025|2026|THOR).)*\.xlsx$")

    print(f"  柜体/台面: {os.path.basename(cab_path)}")
    print(f"  五金:      {os.path.basename(hw_path)}")

    cabinet_pricing, ct_pricing = parse_cabinet_and_countertop(cab_path)
    hardware = parse_hardware(hw_path)
    data = build_data(cabinet_pricing, ct_pricing, hardware)

    sources = [cab_path, hw_path]
    js = to_js(data, sources)
    json5 = to_json5(data, sources)

    with open(OUT_JS, "w", encoding="utf-8", newline="\n") as f:
        f.write(js)
    os.makedirs(os.path.dirname(OUT_JSON5), exist_ok=True)
    with open(OUT_JSON5, "w", encoding="utf-8", newline="\n") as f:
        f.write(json5)

    print(f"[OK] 柜体系列: {len(cabinet_pricing)}")
    print(f"[OK] 台面系列: {len(ct_pricing)}")
    print(f"[OK] 五金条目: {len(hardware)}")
    print(f"[OK] 已写入 {OUT_JS}")
    print(f"[OK] 已写入 {OUT_JSON5}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
